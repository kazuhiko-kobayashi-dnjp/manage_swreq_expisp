#!/usr/bin/env python3
"""
Export requirements.json to Excel in the original BEV要件リスト format.
Usage: python3 export_excel.py <data_json_path> <output_xlsx_path>
"""
import sys
import json
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DATA_FILE = sys.argv[1] if len(sys.argv) > 1 else "data/requirements.json"
OUT_FILE = sys.argv[2] if len(sys.argv) > 2 else "export.xlsx"

with open(DATA_FILE, encoding="utf-8") as f:
    data = json.load(f)

req_data = data.get("req_data", [])
events_cfg = data.get("events_config", [])
events_list = [e["name"] for e in events_cfg] if events_cfg else data.get("events", [])
schedules = data.get("schedules", {})
app_dates = data.get("app_dates", {})

SCHEDULE_LABELS = ["大規模ロジック変更：仕様〆", "小規模ロジック変更：仕様〆", "定数変更：仕様〆", "Vlink 〆", "部集", "PFリリース", "アプリリリース"]
ORIG_SCHED_KEYS  = ["大規模仕様〆", "小規模仕様〆", "定数変更仕様〆", "Vlink〆", "部集", "PFリリース", "アプリリリース"]

STATUS_FILL = {
    "request":   PatternFill("solid", fgColor="FF0000"),
    "agreed":    PatternFill("solid", fgColor="FFFF00"),
    "confirmed": PatternFill("solid", fgColor="CCCCFF"),
}
HEADER_FILL   = PatternFill("solid", fgColor="1a237e")
HEADER2_FILL  = PatternFill("solid", fgColor="283593")
SCHED_FILL    = PatternFill("solid", fgColor="e8eaf6")
HDR_FONT      = Font(color="FFFFFF", bold=True, size=9)
HDR_FONT2     = Font(color="e8eaf6", bold=True, size=8)
DATA_FONT     = Font(size=9)
TINY_FONT     = Font(size=8)
WARN_FONT     = Font(color="FF8B00", bold=True, size=9)
thin          = Side(border_style="thin", color="BDBDBD")
THIN_BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BEV要件織り込みリスト"
ws.sheet_view.showGridLines = True

# =========== FIXED COLUMN WIDTHS ===========
FIXED_COLS = {
    1: 3,   # A
    2: 14,  # B camera
    3: 10,  # C dept
    4: 10,  # D person
    5: 16,  # E 大分類
    6: 16,  # F 中分類
    7: 28,  # G 仕様書
    8: 32,  # H 機能説明
    9: 12,  # I 発行日
    10: 8,  # J SoC
    11: 14, # K SYS3コンポ
    12: 28, # L 仕様書(full)
}
for col_i in range(1, len(events_list) + 37 + 14):
    c = FIXED_COLS.get(col_i, 9)
    ws.column_dimensions[get_column_letter(col_i)].width = c

# Event cols start at col 13
EV_START = 13

# =========== ROW HEIGHTS ===========
for r in range(1, 13):
    ws.row_dimensions[r].height = 22

# =========== HEADER ROW 4: Group labels ===========
ws.row_dimensions[4].height = 20
# Fixed headers
fixed_hdrs = [("B", "C", "カテゴリ"), ("E", "G", "機能分類"), ("H", "H", "機能説明"),
              ("J", "J", "SoC"), ("K", "K", "SYS3コンポ"), ("L", "L", "仕様書")]
def col_num(letter):
    return openpyxl.utils.column_index_from_string(letter)

for start_l, end_l, label in fixed_hdrs:
    sc, ec = col_num(start_l), col_num(end_l)
    ws.cell(4, sc, label).font = HDR_FONT
    ws.cell(4, sc).fill = HEADER_FILL
    ws.cell(4, sc).alignment = CENTER
    if sc != ec:
        ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=ec)

# row 4: event group header
ev_end_col = EV_START + len(events_list) - 1
ws.cell(4, EV_START, "イベント別織り込み状況").font = HDR_FONT
ws.cell(4, EV_START).fill = HEADER_FILL
ws.cell(4, EV_START).alignment = CENTER
if len(events_list) > 1:
    ws.merge_cells(start_row=4, start_column=EV_START, end_row=4, end_column=ev_end_col)

# meta cols after events
META_START = ev_end_col + 1
meta_labels = ["SubTeam", "", "ソフト担当説明", "評価担当説明", "課長承認", "SW変更", "変更概要", "", "要件管理チケット", "", "", "", "管理ID"]
for i, lbl in enumerate(meta_labels):
    c = META_START + i
    cell = ws.cell(4, c, lbl)
    cell.font = HDR_FONT2
    cell.fill = HEADER2_FILL
    cell.alignment = CENTER

# =========== ROW 5: Event names ===========
ws.row_dimensions[5].height = 42
fixed_col5 = {2: "カメラ", 3: "部署", 4: "担当者", 5: "大分類", 6: "中分類",
              7: "小分類(仕様書)", 8: "機能説明", 9: "発行日", 10: "SoC", 11: "SYS3コンポ", 12: "仕様書(フル)"}
for c, lbl in fixed_col5.items():
    cell = ws.cell(5, c, lbl)
    cell.font = HDR_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER

for j, ev in enumerate(events_list):
    c = EV_START + j
    cell = ws.cell(5, c, ev)
    cell.font = HDR_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER

for i, lbl in enumerate(meta_labels):
    c = META_START + i
    ws.cell(5, c, lbl).font = HDR_FONT2
    ws.cell(5, c).fill = HEADER2_FILL
    ws.cell(5, c).alignment = CENTER

# =========== ROWS 6-12: Schedules ===========
for si, (label, key) in enumerate(zip(SCHEDULE_LABELS, ORIG_SCHED_KEYS)):
    row = 6 + si
    ws.row_dimensions[row].height = 18
    ws.cell(row, 8, label).font = Font(bold=True, size=8, color="172b4d")
    ws.cell(row, 8).fill = SCHED_FILL
    ws.cell(row, 8).alignment = LEFT
    sched_data = schedules.get(key, {}) if schedules else {}
    # For events_config deadlines
    for j, ev in enumerate(events_list):
        c = EV_START + j
        # Check events_config deadlines first
        ecfg = next((e for e in events_cfg if e.get("name") == ev), None)
        date_val = (ecfg.get("deadlines", {}).get(key, "") if ecfg else "") or sched_data.get(ev, "")
        if date_val:
            cell = ws.cell(row, c, str(date_val)[:10])
            cell.font = TINY_FONT
            cell.fill = SCHED_FILL
            cell.alignment = CENTER

# =========== DATA ROWS ===========
CAMERA_LABELS = {
    "FCM": "FCMカメラ", "PVM": "PVMカメラ",
    "Side": "Sideカメラ", "SideEIM": "SideEIMカメラ",
    "Tele": "Teleカメラ", "All": "全カメラ"
}

# Compute issue_date → warn threshold
def is_warn(r, ev):
    if not r.get("issue_date"):
        return False
    try:
        im = datetime.strptime(r["issue_date"], "%Y-%m-%d").timestamp()
    except:
        return False
    ecfg = next((e for e in events_cfg if e.get("name") == ev), None)
    for key in ["大規模仕様〆", "小規模仕様〆", "定数変更仕様〆"]:
        date_str = (ecfg.get("deadlines", {}).get(key, "") if ecfg else "") or (schedules.get(key, {}).get(ev, "") if schedules else "")
        if date_str and str(date_str)[:4].isdigit():
            try:
                dl = datetime.strptime(str(date_str)[:10], "%Y-%m-%d").timestamp()
                diff = (dl - im) / 86400
                if -7 <= diff <= 14:
                    return True
            except:
                pass
    return False

JIRA_IDX     = meta_labels.index("要件管理チケット")   # 8
MGMT_ID_IDX  = meta_labels.index("管理ID")             # 12
SW_IDX       = meta_labels.index("SW変更")              # 5
SUMMARY_IDX  = meta_labels.index("変更概要")             # 6
SUBTEAM_IDX  = 0  # SubTeam

data_row = 13
persons_map = {}  # for persons grouping

for r in req_data:
    cameras = r.get("cameras") or ["FCM"]
    for cam in cameras:
        ws.row_dimensions[data_row].height = 18
        # Fixed cols
        ws.cell(data_row, 2, CAMERA_LABELS.get(cam, cam)).font = DATA_FONT
        dept = ""
        person = ""
        persons_list = r.get("persons") or []
        if persons_list:
            person = persons_list[0]
        ws.cell(data_row, 3, dept).font = DATA_FONT
        ws.cell(data_row, 4, person).font = DATA_FONT
        ws.cell(data_row, 5, r.get("cat1","")).font = DATA_FONT
        ws.cell(data_row, 6, r.get("cat2","")).font = DATA_FONT
        ws.cell(data_row, 7, r.get("cat3","")).font = DATA_FONT
        ws.cell(data_row, 8, r.get("desc","")).font = DATA_FONT
        ws.cell(data_row, 8).alignment = LEFT
        ws.cell(data_row, 9, r.get("issue_date","")).font = WARN_FONT if any(is_warn(r, ev) for ev in events_list) else DATA_FONT
        ws.cell(data_row, 10, r.get("soc","")).font = DATA_FONT
        ws.cell(data_row, 11, "").font = DATA_FONT
        ws.cell(data_row, 12, r.get("cat3","")).font = DATA_FONT

        # Event cols
        ep = r.get("event_plan", {})
        for j, ev in enumerate(events_list):
            c = EV_START + j
            cam_data = ep.get(ev, {}).get(cam, {})
            val = cam_data.get("v", "") if cam_data else ""
            st  = cam_data.get("s", "") if cam_data else ""
            if not val or val in ("-", "ー"):
                val = "◎" if st in ("agreed","confirmed") else ("-" if not st else val)
            cell = ws.cell(data_row, c, val if val else "-")
            cell.font = DATA_FONT
            cell.alignment = CENTER
            if st and st in STATUS_FILL:
                cell.fill = STATUS_FILL[st]

        # Meta cols
        ws.cell(data_row, META_START + SUBTEAM_IDX, r.get("subteam","")).font = DATA_FONT
        ws.cell(data_row, META_START + SW_IDX, r.get("sw","")).font = DATA_FONT
        ws.cell(data_row, META_START + SUMMARY_IDX, r.get("summary","")).font = DATA_FONT
        ws.cell(data_row, META_START + SUMMARY_IDX).alignment = LEFT
        jira_str = ", ".join(r.get("jira") or [])
        ws.cell(data_row, META_START + JIRA_IDX, jira_str).font = DATA_FONT
        ws.cell(data_row, META_START + MGMT_ID_IDX, r.get("req_id","")).font = DATA_FONT

        # Border for all cells in row
        for c in range(2, META_START + len(meta_labels)):
            ws.cell(data_row, c).border = THIN_BORDER

        data_row += 1

# Freeze panes after fixed cols + header rows
ws.freeze_panes = ws.cell(13, EV_START)

# Auto-filter on row 5
ws.auto_filter.ref = "B5:" + get_column_letter(META_START + len(meta_labels) - 1) + str(data_row - 1)

wb.save(OUT_FILE)
print(f"Exported {data_row - 13} rows to {OUT_FILE}", file=sys.stderr)
